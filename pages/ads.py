import streamlit as st
import shopee_client as sc
import supabase_client as sdb
import pandas as pd
import io
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BR = ZoneInfo("America/Sao_Paulo")


def _gerar_xlsx_sem_ads(df):
    wb  = Workbook()
    ws  = wb.active
    ws.title = "Produtos sem Ads"
    thin  = Side(style="thin", color="CCCCCC")
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfill = PatternFill("solid", start_color="0F3638")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    yfill = PatternFill("solid", start_color="FFF9C4")

    headers = ["item_id", "nome", "sku", "estoque", "vendas", "preco", "anunciar (sim/nao)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont; cell.fill = hfill
        cell.alignment = Alignment(horizontal="center"); cell.border = brd

    for row_idx, (_, r) in enumerate(df.iterrows(), 2):
        vals = [r.get("item_id",""), r.get("item_name",""), r.get("item_sku",""),
                r.get("stock",0), r.get("sales",0), r.get("price",0), ""]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = brd; cell.font = Font(name="Arial", size=10)
            if col_idx == 7:
                cell.fill = yfill; cell.alignment = Alignment(horizontal="center")
            if col_idx == 6:
                cell.number_format = 'R$ #,##0.00'

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 22

    wi = wb.create_sheet("Instruções")
    rows_i = [
        ["Campo", "Descrição"],
        ["item_id", "Não altere — ID do produto na Shopee"],
        ["anunciar (sim/nao)", "Digite 'sim' para criar Ad, 'nao' para ignorar"],
    ]
    hf2 = PatternFill("solid", start_color="01696F")
    for ri, row in enumerate(rows_i, 1):
        for ci, val in enumerate(row, 1):
            cell = wi.cell(row=ri, column=ci, value=val)
            cell.border = brd
            cell.font = Font(name="Arial", size=10, bold=(ri==1), color="FFFFFF" if ri==1 else "000000")
            if ri == 1: cell.fill = hf2
    wi.column_dimensions["A"].width = 22
    wi.column_dimensions["B"].width = 45

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render():
    st.markdown('<p class="section-title">📣 Anúncios (Ads)</p>', unsafe_allow_html=True)
    st.markdown('<p class="section-sub">Gerencie campanhas GMV Max da Shopee Ads</p>', unsafe_allow_html=True)

    # Saldo rápido
    col_bal, _ = st.columns([2, 4])
    with col_bal:
        if st.button("💰 Ver Saldo de Ads", use_container_width=True):
            with st.spinner("Consultando saldo..."):
                res = sc.get_total_balance()
            if res.get("error"):
                st.error(f"❌ {res['error']}")
            else:
                resp = res.get("response", {})
                paid = resp.get("paid_balance", resp.get("balance", 0))
                free = resp.get("free_balance", 0)
                st.success(f"💳 Saldo pago: **R$ {float(paid):.2f}** | Saldo gratuito: **R$ {float(free):.2f}**")

    st.divider()

    tab_sem, tab_roas, tab_perf = st.tabs([
        "🔍 Produtos SEM Ads",
        "⚡ Criar / Gerenciar Ads",
        "📊 Performance",
    ])

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 1 — PRODUTOS SEM ADS (RECOMENDADOS)
    # ══════════════════════════════════════════════════════════════════════════
    with tab_sem:
        st.markdown("#### Produtos elegíveis para anunciar")
        st.info("A Shopee retorna automaticamente produtos com estoque e sem campanha ativa, classificados como **Hot Search**, **Best Seller** ou **Best ROI**.")

        if st.button("🔍 Carregar Produtos Elegíveis", use_container_width=False):
            with st.spinner("Buscando produtos recomendados..."):
                res = sc.get_all_recommended_items()
            if res.get("error"):
                st.error(f"❌ {res['error']}")
                if res.get("details"):
                    st.json(res["details"])
            else:
                items = res.get("response", {}).get("item_list", [])
                st.session_state["ads_sem"] = items
                if not items:
                    st.info("Nenhum produto elegível. Todos já possuem Ads ou estoque zerado.")
                else:
                    st.success(f"✅ {len(items)} produto(s) elegível(is) para anunciar")

        if st.session_state.get("ads_sem"):
            items = st.session_state["ads_sem"]
            df_sem = pd.DataFrame([{
                "item_id":   it.get("item_id"),
                "item_name": it.get("item_name",""),
                "item_sku":  it.get("item_sku",""),
                "stock":     it.get("stock", 0),
                "sales":     it.get("sales", 0),
                "price":     it.get("price", 0),
                "tag":       it.get("tag",""),
            } for it in items])

            filtro = st.text_input("Filtrar por nome ou SKU", key="filtro_sem_ads")
            df_show = df_sem.copy()
            if filtro.strip():
                df_show = df_show[
                    df_show["item_name"].str.contains(filtro, case=False, na=False) |
                    df_show["item_sku"].str.contains(filtro, case=False, na=False)
                ]

            st.caption(f"{len(df_show)} produto(s)")
            st.dataframe(df_show.rename(columns={
                "item_id":"ID","item_name":"Nome","item_sku":"SKU",
                "stock":"Estoque","sales":"Vendas","price":"Preço (R$)","tag":"Tag Shopee"
            }), use_container_width=True, hide_index=True,
                column_config={"Preço (R$)": st.column_config.NumberColumn(format="R$ %.2f")})

            xlsx_bytes = _gerar_xlsx_sem_ads(df_show)
            st.download_button(
                "📥 Baixar XLSX para selecionar",
                data=xlsx_bytes,
                file_name="produtos_sem_ads.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 2 — CRIAR / GERENCIAR ADS
    # ══════════════════════════════════════════════════════════════════════════
    with tab_roas:
        st.markdown("#### Criar Anúncios via XLSX")
        st.info("""
        **Fluxo:**
        1. Baixe o XLSX na aba **Produtos SEM Ads**
        2. Preencha **sim** na coluna *anunciar* nos produtos desejados
        3. Faça upload aqui e configure ROAS + orçamento
        4. O sistema cria uma campanha **Item GMV Max** por produto
        """)

        arquivo = st.file_uploader("Upload do XLSX preenchido", type=["xlsx","csv"], key="ads_up")
        if arquivo:
            if arquivo.name.endswith(".xlsx"):
                df_up = pd.read_excel(arquivo, dtype={"item_id": str})
            else:
                df_up = pd.read_csv(arquivo, dtype={"item_id": str})

            df_sim = df_up[df_up["anunciar (sim/nao)"].str.strip().str.lower() == "sim"].copy()
            if df_sim.empty:
                st.warning("Nenhum produto marcado como 'sim'.")
            else:
                st.success(f"✅ {len(df_sim)} produto(s) marcados para anunciar")
                st.dataframe(df_sim[["item_id","nome","sku"]].head(20),
                             use_container_width=True, hide_index=True)

                col_r, col_b, col_d = st.columns(3)
                with col_r:
                    roas_alvo = st.number_input("ROAS alvo", 0.1, 100.0, 3.0, 0.1, key="roas_criar")
                with col_b:
                    budget_dia = st.number_input("Orçamento diário (R$)", 1.0, 10000.0, 50.0, 1.0, key="budget_criar")
                with col_d:
                    data_ini = st.date_input("Início", value=datetime.now(BR).date(), key="date_criar")

                start_str = data_ini.strftime("%d-%m-%Y")
                st.caption(f"Serão criadas **{len(df_sim)}** campanhas Item GMV Max com ROAS {roas_alvo} e orçamento R$ {budget_dia:.2f}/dia")

                if st.button("🚀 Criar Anúncios", type="primary", key="criar_ads_btn"):
                    erros   = []
                    sucesso = 0
                    prog    = st.progress(0, text="Criando anúncios...")
                    total   = len(df_sim)

                    for i, (_, row) in enumerate(df_sim.iterrows()):
                        item_id = int(str(row["item_id"]).strip())
                        res = sc.create_manual_product_ads(
                            item_id      = item_id,
                            budget       = budget_dia,
                            roas_target  = roas_alvo,
                            start_date   = start_str,
                            bidding_method = "auto",
                        )
                        if res.get("error"):
                            erros.append(f"ID {item_id}: {res['error']}")
                        else:
                            sucesso += 1
                        prog.progress((i+1)/total, text=f"Criando {i+1} de {total}...")

                    prog.empty()
                    if sucesso:
                        st.success(f"✅ {sucesso} campanha(s) criada(s)!")
                    if erros:
                        st.warning(f"⚠️ {len(erros)} erro(s):")
                        for e in erros[:20]:
                            st.caption(f"• {e}")

        st.divider()
        st.markdown("#### Consultar campanhas de um produto específico")
        item_id_busca = st.text_input("Item ID", placeholder="Ex: 27590274924", key="item_id_camp")
        if st.button("🔍 Buscar Campanhas do Produto", key="buscar_camp"):
            if not item_id_busca.strip():
                st.warning("Informe o Item ID.")
            else:
                with st.spinner("Buscando..."):
                    res = sc.get_product_campaign_id_list(int(item_id_busca.strip()))
                if res.get("error"):
                    st.error(f"❌ {res['error']}")
                else:
                    camp_ids = res.get("response", {}).get("campaign_id_list", [])
                    if not camp_ids:
                        st.info("Nenhuma campanha encontrada para este produto.")
                    else:
                        st.success(f"✅ {len(camp_ids)} campanha(s) encontrada(s)")
                        rows = []
                        for cid in camp_ids:
                            info = sc.get_campaign_setting_info(cid)
                            cfg  = info.get("response", {})
                            rows.append({
                                "Campaign ID":    cid,
                                "Status":         cfg.get("campaign_status",""),
                                "ROAS Alvo":      cfg.get("roas_target", 0),
                                "Orçamento":      cfg.get("budget", 0),
                                "Início":         cfg.get("start_date",""),
                                "Método Bidding": cfg.get("bidding_method",""),
                            })
                        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True,
                                     column_config={
                                         "ROAS Alvo": st.column_config.NumberColumn(format="%.2f"),
                                         "Orçamento": st.column_config.NumberColumn(format="R$ %.2f"),
                                     })

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 3 — PERFORMANCE
    # ══════════════════════════════════════════════════════════════════════════
    with tab_perf:
        st.markdown("#### Performance Geral dos Anúncios CPC")

        col_d1, col_d2 = st.columns(2)
        with col_d1:
            data_ini_p = st.date_input("De", value=(datetime.now(BR) - timedelta(days=7)).date(), key="perf_ini")
        with col_d2:
            data_fim_p = st.date_input("Até", value=datetime.now(BR).date(), key="perf_fim")

        if st.button("📊 Carregar Performance", key="load_perf"):
            with st.spinner("Buscando dados de performance..."):
                res = sc.get_all_cpc_daily_performance(
                    start_date = data_ini_p.strftime("%Y-%m-%d"),
                    end_date   = data_fim_p.strftime("%Y-%m-%d"),
                )
            if res.get("error"):
                st.error(f"❌ {res['error']}")
                if res.get("details"):
                    st.json(res["details"])
            else:
                perf_list = res.get("response", {}).get("daily_performance_list",
                            res.get("response", {}).get("performance_list", []))
                if not perf_list:
                    st.info("Nenhum dado de performance encontrado para o período.")
                    st.json(res)
                else:
                    rows = []
                    for p in perf_list:
                        rows.append({
                            "Data":          p.get("date",""),
                            "Impressões":    p.get("impression", 0),
                            "Cliques":       p.get("click", 0),
                            "CTR (%)":       round(p.get("click_through_rate", 0) * 100, 2),
                            "Gasto (R$)":    p.get("cost", 0),
                            "GMV Ads (R$)":  p.get("gmv", 0),
                            "ROAS Real":     round(p.get("roas", 0), 2),
                            "Conversões":    p.get("order", 0),
                            "CR (%)":        round(p.get("conversion_rate", 0) * 100, 2),
                        })
                    df_perf = pd.DataFrame(rows)

                    # Métricas totais
                    c1, c2, c3, c4 = st.columns(4)
                    c1.metric("Total Gasto",    f"R$ {df_perf['Gasto (R$)'].sum():,.2f}")
                    c2.metric("GMV Gerado",     f"R$ {df_perf['GMV Ads (R$)'].sum():,.2f}")
                    c3.metric("ROAS Médio",     f"{df_perf['ROAS Real'].mean():.2f}")
                    c4.metric("Conversões",     f"{df_perf['Conversões'].sum():,}")

                    st.divider()
                    st.dataframe(df_perf, use_container_width=True, hide_index=True,
                                 column_config={
                                     "Gasto (R$)":   st.column_config.NumberColumn(format="R$ %.2f"),
                                     "GMV Ads (R$)": st.column_config.NumberColumn(format="R$ %.2f"),
                                     "ROAS Real":    st.column_config.NumberColumn(format="%.2f"),
                                 })

                    csv = df_perf.to_csv(index=False).encode("utf-8")
                    st.download_button("📥 Exportar CSV", data=csv,
                                       file_name="performance_ads.csv", mime="text/csv")