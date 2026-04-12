import streamlit as st
import shopee_client as sc
import supabase_client as sdb
import pandas as pd
import io
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

BR = ZoneInfo("America/Sao_Paulo")


def _ts(ts):
    if not ts:
        return ""
    try:
        return datetime.fromtimestamp(int(ts), tz=BR).strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(ts)


def _dt_to_ts(date, hour, minute):
    """Converte date + hora + minuto para timestamp unix em horário de Brasília."""
    dt = datetime(date.year, date.month, date.day, hour, minute, 0, tzinfo=BR)
    return int(dt.timestamp())


def render():
    st.markdown('<p class="section-title">🏷️ Descontos</p>', unsafe_allow_html=True)
    st.markdown('<p class="section-sub">Gerencie promoções — Central de Marketing Shopee</p>', unsafe_allow_html=True)

    st.info("""
    **Fluxo correto:**
    1. **Listar** → veja suas promoções existentes
    2. **Criar** → crie uma nova promoção (nome + período)
    3. **Adicionar Produtos** → adicione produtos com preço promocional (manual ou via CSV)
    """)

    tab_lista, tab_criar, tab_add = st.tabs([
        "📋 Minhas Promoções",
        "➕ Criar Promoção",
        "🛍️ Adicionar Produtos",
    ])

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 1 — LISTAR
    # ══════════════════════════════════════════════════════════════════════════
    with tab_lista:
        col_s, col_b = st.columns([2, 1])
        with col_s:
            disc_status = st.selectbox("Status da promoção",
                ["ongoing", "upcoming", "expired", "paused"],
                format_func=lambda x: {"ongoing":"Em andamento","upcoming":"Agendada",
                                        "expired":"Encerrada","paused":"Pausada"}[x])
        with col_b:
            st.markdown("<br>", unsafe_allow_html=True)
            btn_listar = st.button("🔍 Carregar Promoções", use_container_width=True)

        if btn_listar:
            with st.spinner("Buscando promoções..."):
                result = sc.get_discount_list(status=disc_status)
            if result.get("error"):
                st.error(f"❌ {result['error']}")
                if result.get("details"):
                    st.json(result["details"])
            else:
                discounts = result.get("response", {}).get("discount_list", [])
                st.session_state["discounts"] = discounts
                if not discounts:
                    st.info("Nenhuma promoção encontrada.")
                else:
                    st.success(f"✅ {len(discounts)} promoção(ões) encontrada(s)")

        if st.session_state.get("discounts"):
            rows = []
            for d in st.session_state["discounts"]:
                rows.append({
                    "ID":     d.get("discount_id"),
                    "Nome":   d.get("discount_name", ""),
                    "Status": d.get("discount_status", ""),
                    "Início": _ts(d.get("start_time")),
                    "Fim":    _ts(d.get("end_time")),
                })
            df_d = pd.DataFrame(rows)
            st.dataframe(df_d, use_container_width=True, hide_index=True)

            st.divider()
            st.markdown("#### Ver produtos de uma promoção")
            nomes   = [f"{r['ID']} — {r['Nome']}" for _, r in df_d.iterrows()]
            escolha = st.selectbox("Selecione", nomes, key="disc_sel")
            idx     = nomes.index(escolha)
            disc_id = st.session_state["discounts"][idx]["discount_id"]

            if st.button("🔍 Ver Produtos"):
                with st.spinner("Buscando produtos..."):
                    res = sc.get_discount_items(disc_id)
                if res.get("error"):
                    st.error(f"❌ {res['error']}")
                    st.json(res)
                else:
                    items = res.get("response", {}).get("item_list", [])
                    if items:
                        ri = []
                        for it in items:
                            ri.append({
                                "Item ID":        it.get("item_id"),
                                "Nome":           it.get("item_name",""),
                                "Preço Promo":    it.get("item_promotion_price", 0),
                                "Preço Original": it.get("item_original_price", 0),
                                "Limite/Cliente": it.get("purchase_limit", 0),
                            })
                        st.dataframe(pd.DataFrame(ri), use_container_width=True, hide_index=True,
                                     column_config={
                                         "Preço Promo":    st.column_config.NumberColumn(format="R$ %.2f"),
                                         "Preço Original": st.column_config.NumberColumn(format="R$ %.2f"),
                                     })
                    else:
                        st.info("Nenhum produto nesta promoção ainda.")

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 2 — CRIAR
    # ══════════════════════════════════════════════════════════════════════════
    with tab_criar:
        st.markdown("#### Nova Promoção de Desconto")

        nome_promo = st.text_input("Nome da Promoção *", placeholder="Ex: Wanpy Especial Abril")

        st.markdown("**Período da promoção**")
        col_di, col_hi, col_mi = st.columns(3)
        with col_di:
            data_ini = st.date_input("Data início *", value=datetime.now(BR).date(), key="di")
        with col_hi:
            hora_ini = st.number_input("Hora início", 0, 23,
                                        value=int(st.session_state.get("hora_ini", 9)), key="hi")
        with col_mi:
            min_ini  = st.number_input("Minuto início", 0, 59,
                                        value=int(st.session_state.get("min_ini", 0)), key="mi")

        col_df, col_hf, col_mf = st.columns(3)
        with col_df:
            data_fim = st.date_input("Data fim *",
                                      value=(datetime.now(BR) + timedelta(days=7)).date(), key="df")
        with col_hf:
            hora_fim = st.number_input("Hora fim", 0, 23,
                                        value=int(st.session_state.get("hora_fim", 23)), key="hf")
        with col_mf:
            min_fim  = st.number_input("Minuto fim", 0, 59,
                                        value=int(st.session_state.get("min_fim", 59)), key="mf")

        ts_ini = _dt_to_ts(data_ini, hora_ini, min_ini)
        ts_fim = _dt_to_ts(data_fim, hora_fim, min_fim)

        st.caption(f"Início: {_ts(ts_ini)}  |  Fim: {_ts(ts_fim)}")

        if st.button("✅ Criar Promoção", type="primary"):
            if not nome_promo.strip():
                st.warning("Digite o nome da promoção.")
            elif ts_fim <= ts_ini:
                st.warning("A data de fim deve ser após a data de início.")
            else:
                with st.spinner("Criando promoção..."):
                    res = sc.create_discount(nome_promo.strip(), ts_ini, ts_fim)
                if res.get("error"):
                    st.error(f"❌ {res['error']}")
                    if res.get("details"):
                        st.json(res["details"])
                else:
                    disc_id = res.get("response", {}).get("discount_id")
                    st.success(f"✅ Promoção criada! ID: `{disc_id}`")
                    st.session_state["novo_disc_id"] = disc_id
                    st.info("Agora vá para a aba **🛍️ Adicionar Produtos** para incluir produtos.")

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 3 — ADICIONAR PRODUTOS
    # ══════════════════════════════════════════════════════════════════════════
    with tab_add:
        st.markdown("#### Adicionar Produtos a uma Promoção")

        disc_id_input = st.text_input(
            "ID da Promoção *",
            value=str(st.session_state.get("novo_disc_id", "")),
            placeholder="Cole o ID da promoção"
        )

        tipo_add = st.radio("Como adicionar produtos?",
                            ["Manual (busca e filtra)", "Upload CSV"],
                            horizontal=True)

        # ── VIA CSV ──────────────────────────────────────────────────────────
        if tipo_add == "Upload CSV":
            st.markdown("""
            **Formato do CSV:**
            | item_id | preco_promo | limite |
            |---|---|---|
            | 27590274924 | 29.90 | 0 |
            | 41864392939 | 15.50 | 2 |

            - `item_id`: ID do produto na Shopee
            - `preco_promo`: preço promocional em R$
            - `limite`: limite de unidades por cliente (0 = sem limite)
            """)

            # Template XLSX gerado direto na memória
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                wb = Workbook()
                ws = wb.active
                ws.title = "Desconto"
                headers = ["item_id", "preco_promo", "limite"]
                hf = PatternFill("solid", start_color="0F3638")
                hfont = Font(bold=True, color="FFFFFF", name="Arial")
                thin  = Side(style="thin", color="CCCCCC")
                brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.font = hfont; cell.fill = hf; cell.border = brd
                    cell.alignment = Alignment(horizontal="center")
                for row_idx, row_data in enumerate([[27590274924,29.90,0],[41864392939,15.50,2]], 2):
                    for col_idx, val in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=val)
                        cell.border = brd; cell.font = Font(name="Arial", size=10)
                        if col_idx == 2: cell.number_format = 'R$ #,##0.00'
                ws.column_dimensions["A"].width = 20
                ws.column_dimensions["B"].width = 18
                ws.column_dimensions["C"].width = 20
                import io as _io
                buf = _io.BytesIO()
                wb.save(buf)
                xlsx_bytes = buf.getvalue()
            except Exception as _e:
                xlsx_bytes = b""
                st.warning(f"Não foi possível gerar o template: {_e}")

            st.download_button("📥 Baixar template XLSX", data=xlsx_bytes,
                               file_name="template_desconto.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            arquivo = st.file_uploader("Upload do XLSX ou CSV", type=["xlsx","csv"])
            if arquivo:
                if arquivo.name.endswith(".xlsx"):
                    df_csv = pd.read_excel(arquivo, dtype={"item_id": int})
                else:
                    df_csv = pd.read_csv(arquivo)
                st.dataframe(df_csv, use_container_width=True, hide_index=True)
                st.caption(f"{len(df_csv)} produto(s) no CSV")

                if st.button("🚀 Enviar para Promoção", type="primary", key="csv_send"):
                    if not disc_id_input.strip():
                        st.warning("Informe o ID da promoção.")
                    else:
                        item_list = []
                        for _, row in df_csv.iterrows():
                            item_list.append({
                                "item_id":              int(row["item_id"]),
                                "item_promotion_price": float(row["preco_promo"]),
                                "purchase_limit":       int(row.get("limite", 0)),
                                "model_list":           [],
                            })
                        erros = []
                        prog  = st.progress(0)
                        # Envia em lotes de 100 (limite da API)
                        for i in range(0, len(item_list), 100):
                            lote = item_list[i:i+100]
                            res  = sc.add_discount_items(int(disc_id_input.strip()), lote)
                            if res.get("error"):
                                erros.append(f"Lote {i//100+1}: {res['error']}")
                            prog.progress(min((i+100)/len(item_list), 1.0))
                        prog.empty()
                        if erros:
                            for e in erros:
                                st.error(e)
                        else:
                            st.success(f"✅ {len(item_list)} produto(s) adicionados!")

        # ── MANUAL ───────────────────────────────────────────────────────────
        else:
            col_db2, col_api2, _ = st.columns([1.5, 1.5, 3])
            with col_db2:
                btn_db2 = st.button("⚡ Carregar do Banco", use_container_width=True, key="add_db")
            with col_api2:
                btn_api2 = st.button("🔄 Buscar da API", use_container_width=True, key="add_api")

            if btn_db2:
                df_cat = sdb.carregar_produtos_db()
                if df_cat.empty:
                    st.warning("Banco vazio. Use Buscar da API.")
                else:
                    st.session_state["desconto_catalogo"] = df_cat[df_cat["Status"]=="NORMAL"]
                    st.success(f"✅ {len(st.session_state['desconto_catalogo']):,} produtos!")

            if btn_api2:
                all_ids = sc.get_all_item_ids(status="NORMAL")
                rows = []
                for i in range(0, min(len(all_ids), 500), 50):
                    res = sc.get_item_base_info(all_ids[i:i+50])
                    for it in res.get("response", {}).get("item_list", []):
                        pi = it.get("price_info", [{}])
                        rows.append({
                            "ID":         it.get("item_id"),
                            "Nome":       it.get("item_name",""),
                            "SKU":        it.get("item_sku",""),
                            "Preco (R$)": float(pi[0].get("current_price",0)) if pi else 0,
                            "Status":     "NORMAL",
                        })
                st.session_state["desconto_catalogo"] = pd.DataFrame(rows)
                st.success(f"✅ {len(rows)} produtos carregados!")

            if st.session_state.get("desconto_catalogo") is not None:
                df_cat = st.session_state["desconto_catalogo"]

                col_t, col_b2 = st.columns([1.2, 3])
                with col_t:
                    tipo = st.selectbox("Filtrar por", ["Nome","SKU","ID"], key="add_tipo")
                with col_b2:
                    termo = st.text_input("Termo (use ; para múltiplos)", key="add_termo")

                df_fil = df_cat.copy()
                if termo.strip():
                    termos = [t.strip() for t in termo.split(";") if t.strip()]
                    if tipo == "Nome":
                        mask = pd.Series([False]*len(df_fil), index=df_fil.index)
                        for t in termos:
                            mask |= df_fil["Nome"].str.contains(t, case=False, na=False)
                        df_fil = df_fil[mask]
                    elif tipo == "SKU":
                        mask = pd.Series([False]*len(df_fil), index=df_fil.index)
                        for t in termos:
                            mask |= df_fil["SKU"].str.contains(t, case=False, na=False)
                        df_fil = df_fil[mask]
                    elif tipo == "ID":
                        ids = [int(x) for x in termo.replace(";",",").split(",") if x.strip().isdigit()]
                        df_fil = df_fil[df_fil["ID"].isin(ids)]

                if not df_fil.empty:
                    st.caption(f"{len(df_fil)} produto(s) selecionados")
                    st.dataframe(df_fil[["ID","Nome","SKU","Preco (R$)"]].head(200),
                                 use_container_width=True, hide_index=True,
                                 column_config={"Preco (R$)": st.column_config.NumberColumn(format="R$ %.2f")})

                    st.divider()
                    col_a, col_b, col_c = st.columns(3)
                    with col_a:
                        tipo_desc = st.selectbox("Tipo de desconto", ["Percentual (%)","Valor fixo (R$)"], key="add_td")
                    with col_b:
                        if tipo_desc == "Percentual (%)":
                            pct = st.number_input("Desconto (%)", 1.0, 99.0, 10.0, 0.5, key="add_pct")
                        else:
                            val = st.number_input("Desconto (R$)", 0.01, value=5.0, step=0.01, key="add_val")
                    with col_c:
                        limite = st.number_input("Limite por cliente (0=sem limite)", 0, value=0, key="add_lim")

                    if st.button("🚀 Adicionar à Promoção", type="primary", key="manual_send"):
                        if not disc_id_input.strip():
                            st.warning("Informe o ID da promoção.")
                        else:
                            item_list = []
                            for _, row in df_fil.iterrows():
                                preco = float(row["Preco (R$)"])
                                preco_promo = round(preco*(1-pct/100), 2) if tipo_desc=="Percentual (%)" else round(max(preco-val, 0.01), 2)
                                item_list.append({
                                    "item_id":              int(row["ID"]),
                                    "item_promotion_price": preco_promo,
                                    "purchase_limit":       int(limite),
                                    "model_list":           [],
                                })

                            erros = []
                            prog  = st.progress(0, text="Enviando...")
                            for i in range(0, len(item_list), 100):
                                lote = item_list[i:i+100]
                                res  = sc.add_discount_items(int(disc_id_input.strip()), lote)
                                if res.get("error"):
                                    erros.append(f"Lote {i//100+1}: {res['error']}")
                                prog.progress(min((i+100)/len(item_list), 1.0))
                            prog.empty()
                            if erros:
                                for e in erros:
                                    st.error(e)
                            else:
                                st.success(f"✅ {len(item_list)} produto(s) adicionados à promoção!")